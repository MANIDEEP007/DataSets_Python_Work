'''
Module that prepares 4 Sheets in a Excel workbook
Sheet1 - Indian Data Related to Covid-19 Count
Sheet2 - State Wise Covid-19 Count
Sheet3 - End User's State Covid-19 Count
Sheet4 - Data Visualization
'''

import urllib.request #To get Json Data from URL
import json #To Convert Json to List/Dictionaries
import openpyxl as Excel #To Create workbook
import datetime #To Generate Unique File Names
from openpyxl.styles import Font #To use Various Font Styles

#Timestamp for Unique File generation
TIMESTAMP = datetime.datetime.now().strftime("%d%m%y%H%M%S")


#Get the data from URL and convert to list of dictionaries -State Wise data
COVID_DATA_URL = "https://api.covidindiatracker.com/state_data.json"
COVID_JSON_DATA = json.loads(
    urllib.request.urlopen(COVID_DATA_URL).read(),
    )

#Get the data from URL and convert to dictionary - Total Count in India
COVID_TOTAL_URL = "https://api.covidindiatracker.com/total.json"
COVID_TOTAL_DATA = json.loads(
    urllib.request.urlopen(COVID_TOTAL_URL).read(),
    )

#Get the State of the End User
States = [each_state['state'] for each_state in COVID_JSON_DATA]
State = input("Enter State which is in given list: " + str(States))
print("Populating the Workbook. Please wait some time")

covid_book_obj = Excel.Workbook() #Create Workbook
#Delete default worksheet
del covid_book_obj[covid_book_obj.sheetnames[0]]

#Create Sheets in Workbook
sheet1 = covid_book_obj.create_sheet(title="CountryData")
sheet2 = covid_book_obj.create_sheet(title="StateWise Data")
sheet3 = covid_book_obj.create_sheet(title=State + " Data")
sheet4 = covid_book_obj.create_sheet(title="Data Visualization of Data")

#Font for Keys, Sheet Headings
keys_font = Font(bold=True)
heading_font = Font(bold=True,size=20)

#Populating Sheet1 with necessary Data
sheet1.merge_cells('G2:L2') #Merging Cells for Heading of sheet
sheet1['G2'].value = "INDIA Covid-19 Cases Count"
sheet1['G2'].font = heading_font
sheet1_keys = [
    "confirmed",
    "recovered",
    "active",
    "deaths",
    ]
#Specify the columns to position data in  Sheet1
sheet1_cell_key_col= 8
sheet1_cell_value_col = 9

sheet1.column_dimensions['H'].width = 15 #Set width of Keys column
sheet1_rows = list(range(6,6 + len(sheet1_keys)))
for each_key,index in zip(sheet1_keys,sheet1_rows):
    #Capitalize Key and write
    sheet1.cell(row=index,column=sheet1_cell_key_col).value = \
                                                    each_key.capitalize()
    sheet1.cell(row=index,column=sheet1_cell_key_col).font = keys_font
    sheet1.cell(row=index,column=sheet1_cell_value_col).value = \
                COVID_TOTAL_DATA.get(each_key,"N/A")

#Populating Sheet2 with necessary data
sheet2.merge_cells('D2:L2') #Merging Cells for Heading of sheet
sheet2['D2'].value = "State Wise Covid-19 Details in India"
sheet2['D2'].font = heading_font
sheet2_keys = [
    "state",
    "confirmed",
    "recovered",
    "active",
    "deaths",
    ]
#Write Header
for index in range(len(sheet2_keys)):
    sheet2.cell(row=3,column=index+1).value = \
                                    sheet2_keys[index].capitalize()
    sheet2.cell(row=3,column=index+1).font = keys_font
    #Set width of columns
    sheet2.column_dimensions[chr(index+65)].width = 15
#Write Records
start_row = 4
for each_rec in COVID_JSON_DATA:
    for index in range(len(sheet2_keys)):
        sheet2.cell(row=start_row,column=index+1).value = \
                                            each_rec.get(sheet2_keys[index])
    start_row += 1

#Populate sheet3 with necessary data
if State in States:
    sheet3.merge_cells('D2:L2') #Merging Cells for Heading of sheet
    sheet3['D2'].value = State + " Covid-19 Details in India"
    sheet3['D2'].font = heading_font
    State_Data = []
    for each_rec in COVID_JSON_DATA:
        if each_rec['state'] == State:
            State_Data = each_rec['districtData']
            break
    sheet3_keys = [
        "name",
        "confirmed",
        ]
    #Write Header
    for index in range(len(sheet3_keys)):
        sheet3.cell(row=3,column=index+1).value = \
                                        sheet3_keys[index].capitalize()
        sheet3.cell(row=3,column=index+1).font = keys_font
        #Set width of columns
        sheet3.column_dimensions[chr(index+65)].width = 20
    #Write Records
    start_row = 4
    for each_rec in State_Data:
        for index in range(len(sheet3_keys)):
            sheet3.cell(row=start_row,column=index+1).value = \
                                            each_rec.get(sheet3_keys[index])
        start_row += 1
else:
    print("Entered State Not Exists!!!. So not populating any data in Sheet3")

#Populate sheet4 with necessary data - Chart
#Sheet4 Chart on Sheet2 data - Recovered, Total Cases, Deaths, District
#Selecting Data
chart1_ref_data = Excel.chart.Reference(
    sheet2,
    min_row=3,
    min_col=2,
    max_row=3 + len(COVID_JSON_DATA),
    max_col=len(sheet2_keys),
    )
#Selecting Categories as State Names
chart1_title_ref = Excel.chart.Reference(
    sheet2,
    min_row = 4,
    min_col = 1,
    max_row = 4+ len(COVID_JSON_DATA),
    )
#Creating Chart Object and setting data, properties
chart1_obj = Excel.chart.BarChart()
chart1_obj.title = "Covid-19 Cases and Deaths"
chart1_obj.add_data(
    chart1_ref_data,
    titles_from_data=True,
    )
chart1_obj.y_axis.title = "Cases of Covid-19"
chart1_obj.x_axis.title = "States in India"
chart1_obj.set_categories(chart1_title_ref)
'''
If values are Zero, it may pop dialogue informing
log cannot be applied to those values, Just click Ok
'''
chart1_obj.y_axis.scaling.logBase = 10
sheet4.add_chart(chart1_obj,"A1")
covid_book_obj.save("COVID_" + TIMESTAMP + ".xlsx") #Save Workbook
